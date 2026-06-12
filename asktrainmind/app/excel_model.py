from __future__ import annotations

import re
from dataclasses import dataclass, field
from pathlib import Path
from typing import Any
from urllib.parse import unquote, urlsplit

from openpyxl import load_workbook
from openpyxl.cell.cell import Cell


@dataclass
class DetailRecord:
    title: str
    values: dict[str, str]


@dataclass
class DocumentRecord:
    doc_id: str
    info_doc: str
    config_links: dict[str, str]
    config_link_titles: dict[str, str] = field(default_factory=dict)
    details: list[DetailRecord] = field(default_factory=list)

    def link_title_for_config(self, config_name: str) -> str:
        title = _text(self.config_link_titles.get(config_name, ""))
        if title:
            return title
        return link_title_from_url(self.config_links.get(config_name, ""))


@dataclass
class FunctionRecord:
    id: str
    funzione: str
    tipo: str
    generale_link: str | None
    start_row: int = 0
    end_row: int = 0
    config_names: list[str] = field(default_factory=list)
    documents: list[DocumentRecord] = field(default_factory=list)


class WorkbookParseError(RuntimeError):
    pass


def _text(value: Any) -> str:
    if value is None:
        return ""
    return str(value).strip()


def _formula_text(cell: Cell) -> str:
    value = cell.value
    if hasattr(value, "text"):
        return str(getattr(value, "text", "") or "")
    if isinstance(value, str):
        return value
    if cell.data_type == "f" and isinstance(cell._value, str):
        return cell._value
    return ""


def _extract_link(cell: Cell) -> str:
    link, _ = _extract_link_and_title(cell)
    return link


def link_title_from_url(url: str) -> str:
    clean = _text(url)
    if not clean:
        return ""
    parsed = urlsplit(clean)
    path_like = parsed.path or parsed.fragment or clean
    tail = unquote(path_like.rstrip("/").rsplit("/", maxsplit=1)[-1])
    if tail:
        return tail
    return clean


def _extract_link_and_title(cell: Cell) -> tuple[str, str]:
    doc_id = _text(cell.parent.cell(row=cell.row, column=4).value)
    raw_value = cell.value
    displayed_value = _text(raw_value) if isinstance(raw_value, str) else ""
    hyperlink_display = _text(cell.hyperlink.display) if cell.hyperlink else ""

    if cell.hyperlink and cell.hyperlink.target:
        url = str(cell.hyperlink.target)
        title = _text(cell.hyperlink.display)
        if not title and displayed_value and displayed_value != url and not displayed_value.startswith("="):
            title = displayed_value
        if not title:
            title = link_title_from_url(url)
        return url, title

    formula = _formula_text(cell)
    if not formula:
        url = _text(cell.value)
        return url, link_title_from_url(url)

    m = re.search(r'HYPERLINK\(\s*"([^"]+)"(?:\s*[,;]\s*"([^"]*)")?', formula, flags=re.IGNORECASE)
    if m:
        url = m.group(1)
        friendly = _text(m.group(2))
        if friendly:
            return url, friendly
        if hyperlink_display:
            return url, hyperlink_display
        if displayed_value and displayed_value != url and not displayed_value.startswith("="):
            return url, displayed_value
        return url, link_title_from_url(url)

    m = re.search(r'CONCATENATE\(\s*"([^"]+)"\s*[,;]\s*\$D\d+\s*\)', formula, flags=re.IGNORECASE)
    if m:
        url = f"{m.group(1)}{doc_id}" if doc_id else m.group(1)
        if hyperlink_display:
            return url, hyperlink_display
        if displayed_value and displayed_value != url and not displayed_value.startswith("="):
            return url, displayed_value
        if doc_id:
            return url, doc_id
        return url, link_title_from_url(url)

    if formula.startswith("="):
        return formula, doc_id or formula
    url = _text(cell.value)
    return url, link_title_from_url(url)


def parse_funzioni_sheet(workbook_path: Path | str, sheet_name: str = "Funzioni") -> list[FunctionRecord]:
    wb = load_workbook(filename=workbook_path, data_only=False)
    if sheet_name not in wb.sheetnames:
        raise WorkbookParseError(f"Foglio '{sheet_name}' non trovato")

    ws = wb[sheet_name]
    header_map: dict[str, int] = {}
    for col in range(1, ws.max_column + 1):
        header = _text(ws.cell(1, col).value).lower()
        if header:
            header_map[header] = col

    id_col = header_map.get("id", 1)
    funzione_col = header_map.get("funzione", 2)
    tipo_col = next((c for h, c in header_map.items() if h.startswith("tipo")), 3)
    doc_id_col = next((c for h, c in header_map.items() if h.startswith("doc id")), 4)
    info_col = next((c for h, c in header_map.items() if h.startswith("info doc")), 5)
    generale_col = next((c for h, c in header_map.items() if h.startswith("generale")), ws.max_column)

    config_cols = list(range(info_col + 1, generale_col))
    config_names = {
        col: _text(ws.cell(1, col).value).replace("\n", " / ") or f"CONF_{col}"
        for col in config_cols
    }
    ordered_config_names = [config_names[col] for col in config_cols]

    records: list[FunctionRecord] = []
    current_function: FunctionRecord | None = None
    current_document: DocumentRecord | None = None

    for row in range(2, ws.max_row + 1):
        row_id = _text(ws.cell(row, id_col).value)
        row_funzione = _text(ws.cell(row, funzione_col).value)
        row_tipo = _text(ws.cell(row, tipo_col).value)
        row_doc_id = _text(ws.cell(row, doc_id_col).value)
        row_info = _text(ws.cell(row, info_col).value)

        if row_id and row_funzione:
            if current_function:
                current_function.end_row = max(current_function.end_row, row - 1)
            current_document = None
            current_function = FunctionRecord(
                id=row_id,
                funzione=row_funzione,
                tipo=row_tipo,
                generale_link=_extract_link(ws.cell(row, generale_col)) or None,
                start_row=row,
                end_row=row,
                config_names=ordered_config_names.copy(),
            )
            records.append(current_function)
            continue

        if not current_function:
            continue

        current_function.end_row = row

        if row_doc_id and row_info:
            config_links: dict[str, str] = {}
            config_link_titles: dict[str, str] = {}
            for col in config_cols:
                url, title = _extract_link_and_title(ws.cell(row, col))
                if not url:
                    continue
                config_name = config_names[col]
                config_links[config_name] = url
                if title:
                    config_link_titles[config_name] = title
            current_document = DocumentRecord(
                doc_id=row_doc_id,
                info_doc=row_info,
                config_links=config_links,
                config_link_titles=config_link_titles,
            )
            current_function.documents.append(current_document)
            continue

        if current_document and row_info:
            values = {
                config_names[col]: _text(ws.cell(row, col).value)
                for col in config_cols
                if _text(ws.cell(row, col).value)
            }
            if values:
                current_document.details.append(DetailRecord(title=row_info, values=values))

    return records
