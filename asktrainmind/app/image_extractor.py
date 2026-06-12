from __future__ import annotations

from dataclasses import dataclass
from io import BytesIO
from pathlib import Path

from openpyxl import load_workbook

from asktrainmind.app.excel_model import FunctionRecord


@dataclass
class WorkbookImage:
    row: int
    column: int
    mime_type: str
    data: bytes


def extract_images(workbook_path: str | Path, sheet_name: str = "Funzioni") -> list[WorkbookImage]:
    wb = load_workbook(filename=workbook_path)
    if sheet_name not in wb.sheetnames:
        return []

    ws = wb[sheet_name]
    output: list[WorkbookImage] = []
    for image in ws._images:  # pylint: disable=protected-access
        try:
            row = image.anchor._from.row + 1
            col = image.anchor._from.col + 1
        except Exception:
            row, col = 0, 0

        payload = b""
        mime = "image/png"
        if hasattr(image, "_data"):
            payload = image._data()
        elif hasattr(image, "ref"):
            ref = image.ref
            if isinstance(ref, BytesIO):
                payload = ref.getvalue()
        output.append(WorkbookImage(row=row, column=col, mime_type=mime, data=payload))
    return output


def select_relevant_images(
    records: list[FunctionRecord], images: list[WorkbookImage] | None
) -> tuple[list[WorkbookImage], str | None]:
    if not images:
        return [], None

    ranges = [(record.start_row, record.end_row) for record in records if record.start_row and record.end_row]
    if not ranges:
        return list(images), "Righe sorgente non disponibili: mostro tutte le immagini del workbook."

    selected = []
    for image in images:
        if any(start <= image.row <= end for start, end in ranges):
            selected.append(image)
    return selected, None


# Hook futuro: estendere qui l'estrazione immagini dai documenti SharePoint collegati.
