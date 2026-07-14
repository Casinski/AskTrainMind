from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\__SCRIPTS\DB Flotte ETR1000 Ver_1.7.xlsx")
wb = load_workbook(str(EXCEL_PATH), data_only=True)
ws = wb["Cartelle"]

print("Foglio Cartelle — tutte le righe con valori:")
for row in range(1, ws.max_row + 1):
    row_vals = []
    for col in range(1, ws.max_column + 1):
        val = ws.cell(row, col).value
        if val is not None:
            row_vals.append(f"  [{col}]={str(val)[:60]!r}")
    if row_vals:
        print(f"\nRiga {row}:")
        for v in row_vals:
            print(v)

print("\n\nFormula in D6 (data_only=False):")
wb_f = load_workbook(str(EXCEL_PATH), data_only=False)
ws_f = wb_f["Cartelle"]
for row in range(1, 10):
    for col in range(1, 6):
        val = ws_f.cell(row, col).value
        if val is not None:
            print(f"  Riga {row} Col {col}: {str(val)[:80]!r}")