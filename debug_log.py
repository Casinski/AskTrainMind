"""
debug_scanner_v15_funzai.py
---------------------------
Diagnostica perché excel_scanner non trova celle da compilare
nel file DB Flotte ETR1000 Ver_1.7.xlsx.

Verifica tre cose:
  1. Valori data_only nelle righe "Link" (livello 2)
  2. Riconoscimento corretto dei livelli 1/2/3
  3. Presenza della label "Funzioni AI" e stato delle celle
"""
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\__SCRIPTS\AskTrainMind\DB Flotte ETR1000 Ver_1.7.xlsx")
SHEET = "Funzioni"

print(f"File: {EXCEL_PATH}")
print(f"Esiste: {EXCEL_PATH.exists()}\n")

wb_data  = load_workbook(str(EXCEL_PATH), data_only=True)
wb_write = load_workbook(str(EXCEL_PATH), data_only=False)
ws_data  = wb_data[SHEET]
ws_write = wb_write[SHEET]

# ── 1. Intestazioni riga 1 ────────────────────────────────────────────────
print("=" * 70)
print("INTESTAZIONI RIGA 1")
print("=" * 70)
for col in range(1, ws_write.max_column + 1):
    val = ws_write.cell(1, col).value
    if val is not None:
        print(f"  Col {col}: {val!r}")

# ── 2. Scansione prime 60 righe — tutti e 3 i livelli ─────────────────────
print("\n" + "=" * 70)
print("SCANSIONE RIGHE 2-60 (data_only=False per col A-E, data_only=True per col F+)")
print("=" * 70)

for row in range(2, 61):
    # Leggi dal foglio di scrittura (formule)
    a = ws_write.cell(row, 1).value   # FUNC ID
    b = ws_write.cell(row, 2).value   # DESCRIZIONE
    c = ws_write.cell(row, 3).value   # DOC CARTELLA
    d = ws_write.cell(row, 4).value   # DOC ID
    e = ws_write.cell(row, 5).value   # Info DOC

    # Per i link (col F+) leggi data_only
    link_vals = {}
    for col in range(6, min(ws_write.max_column + 1, 13)):
        v = ws_data.cell(row, col).value
        if v is not None:
            link_vals[col] = str(v)[:25]

    # Determina il tipo di riga
    a_str = str(a).strip() if a and not str(a).startswith("=") else ""
    c_str = str(c).strip() if c else ""
    d_str = str(d).strip() if d and not str(d).startswith("=") else ""
    e_str = str(e).strip() if e else ""

    if not any([a, b, c, d, e] + list(link_vals.values())):
        continue

    row_type = "?"
    if a_str and not c_str:
        row_type = "LIVELLO1"
    elif c_str and d_str and d_str not in ("\\", ""):
        row_type = "LIVELLO2"
    elif e_str:
        row_type = f"LIVELLO3({e_str[:20]})"

    print(f"\n  Riga {row:3d} [{row_type}]:")
    if a is not None: print(f"    A(FUNC_ID)   : {str(a)[:40]!r}")
    if b is not None: print(f"    B(DESC)      : {str(b)[:40]!r}")
    if c is not None: print(f"    C(CARTELLA)  : {c!r}")
    if d is not None: print(f"    D(DOC_ID)    : {str(d)[:40]!r}")
    if e is not None: print(f"    E(Info)      : {e!r}")
    if link_vals:
        print(f"    Link (data_only): {link_vals}")

# ── 3. Cerca tutte le righe "Funzioni AI" nel foglio ─────────────────────
print("\n" + "=" * 70)
print(f"TUTTE LE RIGHE CON 'Funzioni AI' IN COLONNA E (max_row={ws_write.max_row})")
print("=" * 70)

found = 0
for row in range(2, ws_write.max_row + 1):
    e_val = ws_write.cell(row, 5).value
    if e_val and str(e_val).strip() == "Funzioni AI":
        found += 1
        # Controlla le celle di configurazione (F-L) su data_only
        cell_status = {}
        for col in range(6, min(ws_write.max_column + 1, 13)):
            v_data  = ws_data.cell(row, col).value    # valore calcolato (link o testo)
            v_write = ws_write.cell(row, col).value   # formula o testo scritto
            if v_write is not None or v_data is not None:
                cell_status[col] = {
                    "data_only": str(v_data)[:30] if v_data else "None",
                    "formula":   str(v_write)[:30] if v_write else "None",
                }
        print(f"\n  Riga {row}: Funzioni AI trovata")
        for col, status in cell_status.items():
            print(f"    Col {col}: data_only={status['data_only']!r} | formula={status['formula']!r}")

if found == 0:
    print("  ⚠ NESSUNA riga 'Funzioni AI' trovata nel foglio!")
    print("  Verifica che il foglio si chiami esattamente 'Funzioni'")
    print(f"  Fogli presenti: {wb_write.sheetnames}")
else:
    print(f"\n  Totale righe 'Funzioni AI': {found}")

# ── 4. Verifica il nome del foglio ────────────────────────────────────────
print("\n" + "=" * 70)
print("FOGLI PRESENTI NEL WORKBOOK")
print("=" * 70)
print(f"  {wb_write.sheetnames}")