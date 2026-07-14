"""
debug_new_columns.py
--------------------
Diagnostica le nuove colonne C (DOC CARTELLA) e D (DOC ID)
nel foglio Funzioni dopo l'aggiornamento del file Excel.
"""
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\Users\2954534.UTENTI\OneDrive - Gruppo Ferrovie Dello Stato\Lavoro\Gadgets - SW\AskTrainMind\DB Flotte ETR1000 Ver_1.7.xlsx")
SHEET_NAME = "Funzioni"

print("=" * 70)
print("INTESTAZIONI RIGA 1 — tutte le colonne")
print("=" * 70)

wb = load_workbook(str(EXCEL_PATH), data_only=True)
ws = wb[SHEET_NAME]

for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        print(f"  Col {col} ({chr(64+col) if col<=26 else '?'}): {val!r}")

print("\n" + "=" * 70)
print("PRIME 40 RIGHE — colonne A, B, C, D, E e prime 2 config")
print("=" * 70)

for row in range(1, 41):
    a = ws.cell(row, 1).value
    b = ws.cell(row, 2).value
    c = ws.cell(row, 3).value
    d = ws.cell(row, 4).value
    e = ws.cell(row, 5).value
    f = ws.cell(row, 6).value
    g = ws.cell(row, 7).value

    if any(v is not None for v in [a, b, c, d, e, f, g]):
        print(
            f"  Riga {row:3d} | "
            f"A={str(a)[:20]!r:22} | "
            f"B={str(b)[:20]!r:22} | "
            f"C={str(c)[:20]!r:22} | "
            f"D={str(d)[:20]!r:22} | "
            f"E={str(e)[:20]!r:22} | "
            f"F={str(f)[:15]!r:17} | "
            f"G={str(g)[:15]!r:17}"
        )

print("\n" + "=" * 70)
print("FORMULE (data_only=False) — colonne C e D, prime 40 righe")
print("=" * 70)

wb_f = load_workbook(str(EXCEL_PATH), data_only=False)
ws_f = wb_f[SHEET_NAME]

for row in range(1, 41):
    c_val = ws_f.cell(row, 3).value
    d_val = ws_f.cell(row, 4).value
    type_c = type(c_val).__name__
    type_d = type(d_val).__name__

    if c_val is not None or d_val is not None:
        print(f"\n  Riga {row}:")
        print(f"    C ({type_c}): {str(c_val)[:120]!r}")
        print(f"    D ({type_d}): {str(d_val)[:120]!r}")