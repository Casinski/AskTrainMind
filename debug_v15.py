"""
debug_v17.py — diagnostica specifica per il file Ver_1.7_00
Controlla:
  1. Valori data_only nelle celle configurazione (riga doc livello 2)
  2. Struttura colonne CONFIG e MAN_CPR
  3. Corrispondenza nomi configurazione tra i fogli
"""
from pathlib import Path
from openpyxl import load_workbook

EXCEL_PATH = Path(r"C:\__SCRIPTS\AskTrainMind\DB Flotte ETR1000 Ver_1.7_00.xlsx")

wb_data  = load_workbook(str(EXCEL_PATH), data_only=True)
wb_form  = load_workbook(str(EXCEL_PATH), data_only=False)

# ── 1. Prime righe del foglio Funzioni ───────────────────────────────────
print("=" * 70)
print("FOGLIO FUNZIONI — riga 1 (intestazioni)")
print("=" * 70)
ws = wb_data["Funzioni"]
for col in range(1, ws.max_column + 1):
    val = ws.cell(1, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\n" + "=" * 70)
print("FOGLIO FUNZIONI — prime righe 2-50, colonne A-L (data_only)")
print("=" * 70)
for row in range(2, 51):
    row_vals = {}
    for col in range(1, 13):
        val = ws.cell(row, col).value
        if val is not None:
            row_vals[col] = str(val)[:30]
    if row_vals:
        print(f"  Riga {row:3d}: { {k: v for k, v in row_vals.items()} }")

# ── 2. Struttura CONFIG riga 1 ────────────────────────────────────────────
print("\n" + "=" * 70)
print("FOGLIO CONFIG — riga 1 completa (tutte le colonne)")
print("=" * 70)
ws_cfg = wb_data["CONFIG"]
for col in range(1, ws_cfg.max_column + 1):
    val = ws_cfg.cell(1, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\nCONFIG riga 2 (sottointestazioni):")
for col in range(1, ws_cfg.max_column + 1):
    val = ws_cfg.cell(2, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\nCONFIG riga 3 (sottointestazioni):")
for col in range(1, ws_cfg.max_column + 1):
    val = ws_cfg.cell(3, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\nCONFIG prima riga dati (riga 4 o 5):")
for r in [4, 5, 6]:
    for col in range(1, ws_cfg.max_column + 1):
        val = ws_cfg.cell(r, col).value
        if val:
            print(f"  Riga {r} Col {col}: {val!r}")

# ── 3. Struttura MAN_CPR riga 1 ───────────────────────────────────────────
print("\n" + "=" * 70)
print("FOGLIO MAN_CPR — riga 1 completa")
print("=" * 70)
ws_man = wb_data["MAN_CPR"]
for col in range(1, ws_man.max_column + 1):
    val = ws_man.cell(1, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\nMAN_CPR riga 2:")
for col in range(1, ws_man.max_column + 1):
    val = ws_man.cell(2, col).value
    if val:
        print(f"  Col {col}: {val!r}")

print("\nMAN_CPR prime righe dati (3-5):")
for r in [3, 4, 5]:
    for col in range(1, ws_man.max_column + 1):
        val = ws_man.cell(r, col).value
        if val:
            print(f"  Riga {r} Col {col}: {str(val)[:40]!r}")