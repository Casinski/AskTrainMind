"""
funzioni_ai_filter.py
---------------------
Entry point e orchestratore dell'automazione "Funzioni AI".

VERSIONE OTTIMIZZATA:
  - Fase 1 (parsing): regex istantaneo, nessuna LLM
  - Fase 2 (deterministico): regex, nessuna LLM
  - Fase 3-6 (LLM): UNA SOLA chiamata Ollama per gruppo (non per configurazione)
  - Risultato LLM riutilizzato per tutte le configurazioni dello stesso gruppo

Stima tempo: ~1-3 minuti per gruppo invece di ~20 minuti.

COLORAZIONE CELLE:
  Verde  (00AA00): equivalenti
  Rosso  (CC0000): differenze tecniche
  Nero   (000000): unica config OPPURE caso incerto
  Grigio (888888): errore/fallback
"""
from __future__ import annotations

import logging
import sys
import time
from collections import defaultdict
from pathlib import Path

from openpyxl import load_workbook
from openpyxl.styles import Font, Alignment

import config as cfg
import support_loader
import excel_scanner
import document_handler
import ai_synthesizer
from ai_synthesizer import ConfigText, SynthesisResult, synthesize_with_comparison
from function_parser import parse_function_structure, ParsedFunction
from deterministic_comparator import compare_all


# ---------------------------------------------------------------------------
# Colori celle
# ---------------------------------------------------------------------------

FONT_GREEN = Font(color="00AA00", bold=False)
FONT_RED   = Font(color="CC0000", bold=False)
FONT_BLACK = Font(color="000000", bold=False)
FONT_GRAY  = Font(color="888888", bold=False, italic=True)


def _apply_cell_style(cell, result: SynthesisResult) -> None:
    text = result.text
    if text.startswith("[") and text.endswith("]"):
        cell.font      = FONT_GRAY
        cell.alignment = Alignment(wrap_text=True)
        return
    if result.has_differences is None:
        cell.font = FONT_BLACK
    elif result.has_differences is False:
        cell.font = FONT_GREEN
    else:
        cell.font = FONT_RED
    cell.alignment = Alignment(wrap_text=True)


# ---------------------------------------------------------------------------
# Logging
# ---------------------------------------------------------------------------
logging.basicConfig(
    level=logging.INFO,
    format="%(asctime)s [%(levelname)s] %(message)s",
    handlers=[
        logging.FileHandler(
            Path(__file__).parent / "funzioni_ai_filter.log",
            encoding="utf-8",
        ),
        logging.StreamHandler(sys.stdout),
    ],
)
log = logging.getLogger(__name__)


# ---------------------------------------------------------------------------
# Verifica prerequisiti
# ---------------------------------------------------------------------------

def check_prerequisites() -> bool:
    ok = True
    if not ai_synthesizer.check_ollama():
        ok = False
    if not cfg.EXCEL_PATH.exists():
        log.error(f"File Excel non trovato: {cfg.EXCEL_PATH}")
        ok = False
    return ok


# ---------------------------------------------------------------------------
# Salvataggio workbook
# ---------------------------------------------------------------------------

def save_workbook(wb) -> None:
    import time as _time
    max_wait, retry_interval, elapsed = 60, 5, 0
    while elapsed <= max_wait:
        try:
            wb.save(str(cfg.EXCEL_PATH))
            log.info(f"✅ File Excel salvato: {cfg.EXCEL_PATH}")
            return
        except PermissionError:
            if elapsed == 0:
                log.warning(f"⚠ File bloccato. Riprovo ogni {retry_interval}s...")
            _time.sleep(retry_interval)
            elapsed += retry_interval

    backup = cfg.EXCEL_PATH.with_name(
        cfg.EXCEL_PATH.stem + "_backup" + cfg.EXCEL_PATH.suffix
    )
    try:
        wb.save(str(backup))
        log.warning(f"⚠ Timeout. Backup salvato: {backup}")
    except Exception as exc:
        log.error(f"Errore salvataggio backup: {exc}")


# ---------------------------------------------------------------------------
# Orchestrazione principale
# ---------------------------------------------------------------------------

def run(start_from_func_id: str = "") -> int:
    start = start_from_func_id.strip() or cfg.START_FROM_FUNC_ID.strip()
    if start:
        log.info(f"⏩ Partenza da funzione: '{start}'")

    # ── Apri workbook ─────────────────────────────────────────────────────
    try:
        wb_data = load_workbook(str(cfg.EXCEL_PATH), data_only=True)
        ws_data = wb_data[cfg.SHEET_FUNZIONI]
    except Exception as exc:
        log.error(f"Impossibile aprire il file (data_only): {exc}")
        return 0

    try:
        wb_write = load_workbook(str(cfg.EXCEL_PATH), data_only=False)
        ws_write = wb_write[cfg.SHEET_FUNZIONI]
    except PermissionError:
        log.error("Il file Excel è aperto. Chiudilo prima di eseguire lo script.")
        return 0
    except Exception as exc:
        log.error(f"Impossibile aprire il file (write): {exc}")
        return 0

    # ── Fogli di supporto ─────────────────────────────────────────────────
    sd = support_loader.load(wb_data)
    if not sd.is_valid():
        log.error("Fogli di supporto non validi (base URL mancante).")
        return 0

    # ── FillTarget ────────────────────────────────────────────────────────
    all_targets = list(
        excel_scanner.scan(ws_write, ws_data, sd, start_from_func_id=start)
    )
    if not all_targets:
        log.info("Nessuna cella da compilare.")
        return 0

    groups: dict[tuple[str, str], list] = defaultdict(list)
    for target in all_targets:
        groups[(target.func_id, target.doc_id)].append(target)

    log.info(
        f"\nFillTarget: {len(all_targets)} celle "
        f"in {len(groups)} gruppi"
    )

    # Svuota cache LLM all'avvio di ogni run
    ai_synthesizer._group_cache.clear()

    filled_count = 0

    for (func_id, doc_id), group_targets in groups.items():

        log.info(
            f"\n{'═'*55}\n"
            f"  {func_id} / {doc_id} "
            f"({len(group_targets)} config)\n"
            f"{'═'*55}"
        )

        # ── Estrazione testi ──────────────────────────────────────────────
        config_texts: list[ConfigText] = []
        for target in group_targets:
            log.info(f"  [{target.config_name}] Estrazione pag.{target.page_number}...")
            doc_path = document_handler.download(target.url)
            if not doc_path:
                config_texts.append(ConfigText(
                    config_name=target.config_name,
                    page_number=target.page_number,
                    text="",
                ))
                log.warning(f"  [{target.config_name}] Documento non trovato")
                continue

            page_text, page_end, start_partial, end_partial = (
                document_handler.extract_page_text(
                    doc_path, target.page_number,
                    func_desc=target.func_desc,
                    func_id=target.func_id,
                )
            )
            target.page_number_end  = page_end
            target.start_is_partial = start_partial
            target.end_is_partial   = end_partial

            config_texts.append(ConfigText(
                config_name=target.config_name,
                page_number=target.page_number,
                page_number_end=page_end,
                text=page_text,
            ))

        valid_texts = [ct for ct in config_texts if ct.text.strip()]
        log.info(f"  Testi: {len(valid_texts)}/{len(group_targets)}")

        # ── Fase 1: parsing regex (istantaneo, nessuna LLM) ──────────────
        parsed_list: list[ParsedFunction] = []
        if len(valid_texts) > 1:
            for ct in valid_texts:
                pf = parse_function_structure(
                    config_name=ct.config_name,
                    text=ct.text,
                    func_id=func_id,
                    func_desc=group_targets[0].func_desc,
                )
                parsed_list.append(pf)

        # ── Fase 2: confronto deterministico (istantaneo, nessuna LLM) ───
        det_report: dict = {
            "any_objective_differences": False,
            "all_differences_summary": [],
        }
        if len(parsed_list) > 1:
            det_report = compare_all(parsed_list)
            n_diffs = len(det_report.get("all_differences_summary", []))
            if det_report["any_objective_differences"]:
                log.info(f"  ⚡ Diff. oggettive: {n_diffs} voci")
            else:
                log.info("  ✅ Nessuna diff. oggettiva")

        # ── Fasi 3-6: LLM — UNA SOLA chiamata per il gruppo ─────────────
        # La chiamata viene fatta PRIMA del ciclo sulle configurazioni.
        # synthesize_with_comparison userà la cache per le config successive.
        if len(valid_texts) > 1:
            log.info(
                f"  Chiamata Ollama (unica per questo gruppo, "
                f"{len(valid_texts)} config in parallelo nel prompt)..."
            )

        for target in group_targets:
            my_text = next(
                (ct.text for ct in config_texts
                 if ct.config_name == target.config_name),
                "",
            )
            cell = ws_write.cell(target.row, target.col)

            if not my_text:
                err = SynthesisResult(
                    text=f"[Testo non estratto: pag.{target.page_number} di {target.doc_id}]",
                    has_differences=None,
                )
                cell.value = err.text
                _apply_cell_style(cell, err)
                filled_count += 1
                continue

            result: SynthesisResult = synthesize_with_comparison(
                func_id=target.func_id,
                func_desc=target.func_desc,
                doc_id=target.doc_id,
                config_name=target.config_name,
                page_number=target.page_number,
                page_text=my_text,
                all_config_texts=valid_texts,
                parsed_list=parsed_list,
                det_report=det_report,
            )

            # ── Nota pagina ───────────────────────────────────────────────
            page_end    = target.page_number_end or target.page_number
            start_label = (
                f"parte di pag.{target.page_number}"
                if getattr(target, "start_is_partial", False)
                else f"pag.{target.page_number}"
            )
            end_label = (
                f"parte di pag.{page_end}"
                if getattr(target, "end_is_partial", False)
                else f"pag.{page_end}"
            )
            page_nota = (
                f"\nPag. iniziale: {start_label} — Pag. finale: {end_label}."
                if page_end > target.page_number
                else f"\nPag. di riferimento: {start_label}."
            )

            final_result = SynthesisResult(
                text=result.text + page_nota,
                has_differences=result.has_differences,
                uncertain=result.uncertain,
                checklist=result.checklist,
                score=result.score,
                technical_differences=result.technical_differences,
                editorial_differences=result.editorial_differences,
                det_summary=result.det_summary,
            )

            cell.value = final_result.text
            _apply_cell_style(cell, final_result)
            filled_count += 1

            color_label = (
                "🟢 VERDE"           if result.has_differences is False else
                "🔴 ROSSO"           if result.has_differences is True  else
                "⚫ NERO (incerto)"  if result.uncertain               else
                "⚫ NERO"
            )
            log.info(
                f"  [{target.config_name}] {color_label} | "
                f"Score: {result.score}"
            )

            # Pausa SOLO dopo la prima config del gruppo
            # (le altre usano la cache, nessuna chiamata LLM aggiuntiva)
            if target == group_targets[0]:
                time.sleep(cfg.AI_CALL_DELAY_SECONDS)

            if cfg.MAX_CELLS_PER_RUN > 0 and filled_count >= cfg.MAX_CELLS_PER_RUN:
                log.info(f"Limite MAX_CELLS_PER_RUN raggiunto ({cfg.MAX_CELLS_PER_RUN}).")
                save_workbook(wb_write)
                return filled_count

    log.info(f"\n{'='*60}\n  CELLE COMPILATE: {filled_count}\n{'='*60}")
    if filled_count > 0:
        save_workbook(wb_write)
    else:
        log.info("Nessuna cella compilata.")

    return filled_count


def main() -> None:
    import argparse
    parser = argparse.ArgumentParser(
        description="Compila le celle 'Funzioni AI' nel foglio Excel ETR1000.",
    )
    parser.add_argument("--start", "-s", metavar="FUNC_ID", default="",
                        help="FUNC ID della funzione da cui iniziare.")
    args = parser.parse_args()

    log.info("\n" + "=" * 60)
    log.info("  funzioni_ai_filter — avvio (versione ottimizzata)")
    log.info("=" * 60)
    log.info(f"  Excel     : {cfg.EXCEL_PATH}")
    log.info(f"  AI model  : Ollama / {cfg.OLLAMA_MODEL}")

    start = args.start.strip() or cfg.START_FROM_FUNC_ID.strip()
    log.info(f"  Partenza  : {'funzione ' + repr(start) if start else 'inizio foglio'}")

    if not check_prerequisites():
        sys.exit(1)

    run(start_from_func_id=start)
    log.info("Fine.")


if __name__ == "__main__":
    main()