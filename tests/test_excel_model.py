from pathlib import Path

from openpyxl import Workbook

from asktrainmind.app.excel_model import link_title_from_url, parse_funzioni_sheet


ROOT = Path(__file__).resolve().parents[1]
WORKBOOK = ROOT / "DB Flotte ETR1000 Ver_0.5_MM.xlsx"


def test_parse_funzioni_sheet_found_and_level1_fields():
    records = parse_funzioni_sheet(WORKBOOK)
    assert records
    first = records[0]
    assert first.id
    assert first.funzione


def test_parse_doc_and_detail_rows_for_fs_dm1():
    records = parse_funzioni_sheet(WORKBOOK)
    fam = next(r for r in records if r.id == "Isolam_FAM_01")
    fs_dm1 = next(d for d in fam.documents if d.doc_id == "FS_DM1")

    assert any("VZI_IT_Base" in cfg for cfg in fs_dm1.config_links)
    assert any("VZI-6_IT_R1" in cfg for cfg in fs_dm1.config_links)
    assert not any("VZ-FR" in cfg for cfg in fs_dm1.config_links)

    titles = {d.title for d in fs_dm1.details}
    assert "Rif. Pagina" in titles
    assert "Componenti circuito elettrico" in titles


def test_config_link_title_from_workbook_row():
    records = parse_funzioni_sheet(WORKBOOK)
    fam = next(r for r in records if r.id == "Isolam_FAM_01")
    fs_dm1 = next(d for d in fam.documents if d.doc_id == "FS_DM1")
    cfg = "VZI_IT_Base / VZI-50_ IT Flotta base"
    assert fs_dm1.link_title_for_config(cfg) == "FS_DM1"


def test_parse_hyperlink_friendly_name_and_concatenate_title(tmp_path):
    wb = Workbook()
    ws = wb.active
    ws.title = "Funzioni"
    headers = [
        "ID",
        "FUNZIONE",
        "TIPO (TBD)",
        "DOC ID",
        "Info DOC",
        "CONF_A",
        "CONF_B",
        "Generale",
    ]
    for col, header in enumerate(headers, start=1):
        ws.cell(row=1, column=col, value=header)

    ws.cell(row=2, column=1, value="ID_X")
    ws.cell(row=2, column=2, value="Funzione X")
    ws.cell(row=2, column=3, value="TIPO")

    ws.cell(row=3, column=4, value="DOC_77")
    ws.cell(row=3, column=5, value="Documento test")
    ws.cell(row=3, column=6, value='=HYPERLINK("https://example.com/spec/abc.pdf","Manuale ABC")')
    ws.cell(row=3, column=7, value='=CONCATENATE("https://example.com/docs/",$D3)')

    path = tmp_path / "titles.xlsx"
    wb.save(path)

    records = parse_funzioni_sheet(path)
    doc = records[0].documents[0]
    assert doc.link_title_for_config("CONF_A") == "Manuale ABC"
    assert doc.link_title_for_config("CONF_B") == "DOC_77"
    assert link_title_from_url("https://example.com/path/demo%20file.pdf?x=1") == "demo file.pdf"
